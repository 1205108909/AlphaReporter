#!/usr/bin/env python
# encoding: utf-8

"""
@Author : wangzhaoyun
@Contact:1205108909@qq.com
@File : ExcelHelper.py 
@Time : 2020/8/6 10:43 
"""
import openpyxl
import pandas as pd


class ExcelHelper(object):
    _dictSheetStartrow = {}
    _dictSheetLastdfShape = {}

    @classmethod
    def createExcel(cls, path):
        wb = openpyxl.Workbook()
        wb.save(path)

    @classmethod
    def Append_df_to_excel(cls, df, file_name, sheet_name, interval=0, sep_key='file_name',
                           truncate_sheet=False,
                           **to_excel_kwargs):

        if sep_key == 'all_name':
            key = file_name + sheet_name
        elif sep_key == 'file_name':
            key = file_name
        else:
            key = sheet_name
        writer = pd.ExcelWriter(file_name, engine='openpyxl')

        try:
            if not key in cls._dictSheetStartrow.keys():
                cls._dictSheetStartrow[key] = 0

            if not key in cls._dictSheetLastdfShape.keys():
                cls._dictSheetLastdfShape[key] = 0

            writer.book = openpyxl.load_workbook(file_name)

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
            cls._dictSheetStartrow[key] += cls._dictSheetLastdfShape[key] + interval
            df.to_excel(writer, sheet_name, startrow=cls._dictSheetStartrow[key], index=False, **to_excel_kwargs)
            cls._dictSheetLastdfShape[key] = df.shape[0]
            writer.close()

        except Exception as e:
            print(f'Append_df_to_excel is Fail')
            print(e)
        else:
            writer.save()

    @classmethod
    def removeSheet(cls, path, sheetName):
        wb = openpyxl.load_workbook(path)
        ws = wb[sheetName]
        wb.remove(ws)
        wb.save(path)
